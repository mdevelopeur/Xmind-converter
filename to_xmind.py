import os
import json
import zipfile
import uuid
from openpyxl import load_workbook

def create_topic(title):
    """Генерирует узел (топик) с уникальным ID"""
    return {
        "id": str(uuid.uuid4()),
        "title": str(title) if title is not None else "",
        "children": {"attached": []}
    }

def build_tree_from_xlsx(xlsx_path):
    """Читает Excel и строит древовидную JSON-структуру для XMind"""
    wb = load_workbook(xlsx_path, data_only=True)
    sheet = wb.active

    root_topic = None
    # Хранилище последних активных узлов для каждого уровня вложенности (индекса колонки)
    last_nodes_at_level = {}

    for row in sheet.iter_rows(values_only=True):
        # Удаляем пустые ячейки в конце строки
        row_cells = [cell for cell in row if cell is not None or row.index(cell) < len(row)]
        
        for col_idx, cell_value in enumerate(row):
            if cell_value is None or str(cell_value).strip() == "":
                continue

            # Создаем новый элемент карты
            new_node = create_topic(cell_value)

            if col_idx == 0:
                # Первая колонка — это всегда центральный корень всей карты
                if root_topic is None:
                    root_topic = new_node
                    last_nodes_at_level[0] = root_topic
            else:
                # Ищем родителя на предыдущем уровне (col_idx - 1)
                parent_level = col_idx - 1
                while parent_level >= 0 and parent_level not in last_nodes_at_level:
                    parent_level -= 1
                
                if parent_level >= 0:
                    parent_node = last_nodes_at_level[parent_level]
                    parent_node["children"]["attached"].append(new_node)
                    last_nodes_at_level[col_idx] = new_node

            # Очищаем память о более глубоких уровнях, так как мы зашли на новую ветку
            levels_to_del = [lvl for lvl in last_nodes_at_level if lvl > col_idx]
            for lvl in levels_to_del:
                del last_nodes_at_level[lvl]

    return root_topic

def save_as_xmind(root_topic, output_xmind_path):
    """Упаковывает JSON-структуру в валидный .xmind архив"""
    # Минимально необходимая структура метаданных XMind
    manifest_data = {
        "file-entries": {
            "content.json": {"content-type": "application/json"},
            "metadata.json": {"content-type": "application/json"}
        }
    }
    
    metadata_data = {
        "creator": {"name": "Python XLSX to XMind Converter", "version": "1.0"}
    }

    # Корневой контейнер листа XMind
    content_data = [{
        "id": str(uuid.uuid4()),
        "class": "sheet",
        "title": "Sheet 1",
        "rootTopic": root_topic
    }]

    # Создаем ZIP-архив с расширением .xmind
    with zipfile.ZipFile(output_xmind_path, 'w', zipfile.ZIP_DEFLATED) as xmind_zip:
        xmind_zip.writestr("manifest.json", json.dumps(manifest_data, ensure_ascii=False, indent=2))
        xmind_zip.writestr("metadata.json", json.dumps(metadata_data, ensure_ascii=False, indent=2))
        xmind_zip.writestr("content.json", json.dumps(content_data, ensure_ascii=False, indent=2))

    print(f"Файл успешно создан: {output_xmind_path}")

def xlsx_to_xmind(xlsx_path, xmind_path):
    if not os.path.exists(xlsx_path):
        print(f"Ошибка: Файл {xlsx_path} не найден.")
        return
        
    print("Анализ Excel и сборка дерева...")
    root = build_tree_from_xlsx(xlsx_path)
    if root:
        save_as_xmind(root, xmind_path)
    else:
        print("Не удалось прочитать данные или файл Excel пуст.")

# --- Пример использования ---
if __name__ == "__main__":
    input_excel = "to_xmind.xlsx"    # Ваша таблица Excel
    output_xmind = "from_excel.xmind" # Итоговая карта XMind
    
    xlsx_to_xmind(input_excel, output_xmind)
